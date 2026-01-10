"""
POD Status Tracker - /pod-status
Consolidate POD status and track closure readiness.

Usage:
    python pod_status.py [master_excel] [--check-report path] [--issues-report path]
"""
import sys
import argparse
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional

import pandas as pd

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))
from shared.config import PODConfig
from shared.excel_utils import write_report, apply_status_formatting


def load_check_report(report_path: Path) -> Dict[str, str]:
    """
    Load POD check report and extract presence status.

    Args:
        report_path: Path to pod-check report

    Returns:
        Dict mapping delivery_id to presence status
    """
    if not report_path or not report_path.exists():
        return {}

    # Find the header row dynamically
    df_raw = pd.read_excel(report_path, header=None)
    header_row = 0
    for i, row in df_raw.iterrows():
        if 'Delivery ID' in str(row.values):
            header_row = i
            break

    df = pd.read_excel(report_path, skiprows=header_row)
    presence = {}

    for _, row in df.iterrows():
        delivery_id = str(row.get('Delivery ID', '')).strip()
        status = str(row.get('Status', '')).strip()
        if delivery_id and delivery_id != 'nan':
            presence[delivery_id] = status

    return presence


def load_issues_report(report_path: Path) -> Dict[str, Dict]:
    """
    Load POD issues report and extract issue details.

    Args:
        report_path: Path to pod-issues report

    Returns:
        Dict mapping delivery_id to issue details
    """
    if not report_path or not report_path.exists():
        return {}

    # Find the header row dynamically
    df_raw = pd.read_excel(report_path, header=None)
    header_row = 0
    for i, row in df_raw.iterrows():
        if 'Delivery ID' in str(row.values):
            header_row = i
            break

    df = pd.read_excel(report_path, skiprows=header_row)
    issues = {}

    for _, row in df.iterrows():
        delivery_id = str(row.get('Delivery ID', '')).strip()
        if delivery_id and delivery_id != 'nan':
            # Only track high/medium severity issues
            severity = str(row.get('Severity', '')).strip()
            if severity in ['High', 'Medium']:
                issues[delivery_id] = {
                    'issue_type': row.get('Issue Type', ''),
                    'severity': severity,
                    'details': row.get('Details', ''),
                }

    return issues


def consolidate_status(
    master_df: pd.DataFrame,
    presence: Dict[str, str],
    issues: Dict[str, Dict]
) -> pd.DataFrame:
    """
    Consolidate status from all sources.

    Args:
        master_df: Master Excel DataFrame
        presence: Dict from pod-check report
        issues: Dict from pod-issues report

    Returns:
        Consolidated status DataFrame
    """
    rows = []

    for _, row in master_df.iterrows():
        delivery_id = str(row.get('delivery_id', '')).strip()

        # Determine presence
        if delivery_id in presence:
            pod_received = 'Yes' if presence[delivery_id] == 'Present' else 'No'
        else:
            pod_received = 'Unknown'

        # Check for issues
        issue_info = issues.get(delivery_id, {})
        has_issues = 'Yes' if issue_info else 'No'
        issue_details = issue_info.get('details', '') if issue_info else ''
        severity = issue_info.get('severity', '')

        # Determine resolution status
        current_status = str(row.get('status', '')).strip().lower()
        if current_status in ['closed', 'complete', 'resolved']:
            resolution_status = 'Closed'
        elif has_issues == 'Yes':
            resolution_status = 'Has Issues'
        elif pod_received == 'No':
            resolution_status = 'Pending POD'
        elif pod_received == 'Yes':
            resolution_status = 'Ready to Close'
        else:
            resolution_status = 'Unknown'

        # Ready to close?
        ready_to_close = 'Yes' if (
            pod_received == 'Yes' and
            has_issues == 'No' and
            resolution_status != 'Closed'
        ) else 'No'

        rows.append({
            'Delivery ID': delivery_id,
            'Customer': row.get('customer', ''),
            'Delivery Date': row.get('date', ''),
            'POD Received': pod_received,
            'Has Issues': has_issues,
            'Issue Details': issue_details,
            'Severity': severity,
            'Resolution Status': resolution_status,
            'Ready to Close': ready_to_close,
        })

    df = pd.DataFrame(rows)

    # Sort: Ready to close first, then Has Issues, then Pending
    status_order = {
        'Ready to Close': 0,
        'Has Issues': 1,
        'Pending POD': 2,
        'Closed': 3,
        'Unknown': 4
    }
    df['_sort'] = df['Resolution Status'].map(status_order)
    df = df.sort_values(['_sort', 'Delivery ID']).drop('_sort', axis=1)

    return df


def run_pod_status(
    master_path: str = None,
    check_report: str = None,
    issues_report: str = None,
    output_folder: str = None
) -> Path:
    """
    Main function to run POD status consolidation.

    Args:
        master_path: Path to master Excel
        check_report: Path to pod-check report
        issues_report: Path to pod-issues report
        output_folder: Path for output report

    Returns:
        Path to generated report
    """
    # Initialize config
    config = PODConfig(
        manifest_path=master_path,
        output_folder=output_folder
    )

    # Use manifest path as master if not provided separately
    master_file = Path(master_path) if master_path else config.manifest_path

    if not master_file.exists():
        print(f"Error: Master file not found: {master_file}")
        sys.exit(1)

    print(f"Master Excel: {master_file}")
    print("-" * 50)

    # Read master Excel
    print("Reading master data...")
    master_df = pd.read_excel(master_file)

    # Standardize column names
    column_map = {v: k for k, v in PODConfig.MANIFEST_COLUMNS.items() if v in master_df.columns}
    master_df = master_df.rename(columns=column_map)

    if 'delivery_id' in master_df.columns:
        master_df['delivery_id'] = master_df['delivery_id'].astype(str).str.strip()

    print(f"Found {len(master_df)} entries in master")

    # Load additional reports
    presence = {}
    issues = {}

    if check_report:
        print(f"Loading check report: {check_report}")
        presence = load_check_report(Path(check_report))
        print(f"  - {len(presence)} presence records loaded")

    if issues_report:
        print(f"Loading issues report: {issues_report}")
        issues = load_issues_report(Path(issues_report))
        print(f"  - {len(issues)} issue records loaded")

    # Consolidate status
    print("Consolidating status...")
    status_df = consolidate_status(master_df, presence, issues)

    # Calculate summary
    summary = {
        'Total Deliveries': len(status_df),
        'PODs Received': len(status_df[status_df['POD Received'] == 'Yes']),
        'PODs Missing': len(status_df[status_df['POD Received'] == 'No']),
        'Has Issues': len(status_df[status_df['Has Issues'] == 'Yes']),
        'Ready to Close': len(status_df[status_df['Ready to Close'] == 'Yes']),
        'Already Closed': len(status_df[status_df['Resolution Status'] == 'Closed']),
        'Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

    # Print summary
    print("-" * 50)
    print("SUMMARY")
    print("-" * 50)
    for key, value in summary.items():
        print(f"{key}: {value}")

    # Write report
    output_path = config.get_output_path('pod_status_report')
    write_report(status_df, output_path, 'POD Status', summary)

    # Apply formatting
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    wb = load_workbook(output_path)
    ws = wb.active

    # Custom formatting for Ready to Close
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in ws.iter_rows(min_row=9, max_row=ws.max_row):
        ready_cell = row[8]  # Ready to Close column
        status_cell = row[7]  # Resolution Status column

        if ready_cell.value == 'Yes':
            for cell in row:
                cell.fill = green_fill
        elif status_cell.value == 'Has Issues':
            for cell in row:
                cell.fill = red_fill
        elif status_cell.value == 'Pending POD':
            for cell in row:
                cell.fill = yellow_fill

    wb.save(output_path)

    print("-" * 50)
    print(f"Report saved: {output_path}")

    # Print closure-ready list
    ready_list = status_df[status_df['Ready to Close'] == 'Yes']['Delivery ID'].tolist()
    if ready_list:
        print(f"\nReady to close ({len(ready_list)} items):")
        for did in ready_list[:10]:
            print(f"  - {did}")
        if len(ready_list) > 10:
            print(f"  ... and {len(ready_list) - 10} more")

    return output_path


def main():
    """Parse arguments and run POD status."""
    parser = argparse.ArgumentParser(
        description='POD Status Tracker - Consolidate status and track closure'
    )
    parser.add_argument(
        'master_excel',
        nargs='?',
        help='Path to master tracking Excel'
    )
    parser.add_argument(
        '--check-report', '-c',
        help='Path to pod-check report'
    )
    parser.add_argument(
        '--issues-report', '-i',
        help='Path to pod-issues report'
    )
    parser.add_argument(
        '--output', '-o',
        help='Output folder for report'
    )

    args = parser.parse_args()

    run_pod_status(
        master_path=args.master_excel,
        check_report=args.check_report,
        issues_report=args.issues_report,
        output_folder=args.output
    )


if __name__ == '__main__':
    main()
