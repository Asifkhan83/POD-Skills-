"""
POD Presence Validator - /pod-check
Compare scanned POD PDFs against manifest to find missing/extra PODs.

Usage:
    python pod_check.py [pod_folder] [manifest_path] [--output output_folder]
"""
import sys
import argparse
from pathlib import Path
from datetime import datetime
from typing import Set, Dict, List, Tuple

import pandas as pd

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))
from shared.config import PODConfig, parse_delivery_id
from shared.excel_utils import read_manifest, write_report, apply_status_formatting


def scan_pod_folder(folder: Path, extensions: List[str]) -> Dict[str, Path]:
    """
    Scan folder for POD files and extract delivery IDs.

    Args:
        folder: Path to POD folder
        extensions: List of valid file extensions

    Returns:
        Dict mapping delivery_id to file path
    """
    pod_files = {}

    for ext in extensions:
        for file_path in folder.glob(f"*{ext}"):
            delivery_id = parse_delivery_id(file_path.name)
            if delivery_id:
                pod_files[delivery_id] = file_path

    return pod_files


def compare_pods(
    manifest_ids: Set[str],
    scanned_ids: Set[str]
) -> Tuple[Set[str], Set[str], Set[str]]:
    """
    Compare manifest IDs with scanned POD IDs.

    Args:
        manifest_ids: Set of delivery IDs from manifest
        scanned_ids: Set of delivery IDs from scanned files

    Returns:
        Tuple of (present, missing, extra) sets
    """
    present = manifest_ids & scanned_ids
    missing = manifest_ids - scanned_ids
    extra = scanned_ids - manifest_ids

    return present, missing, extra


def create_report_dataframe(
    manifest_df: pd.DataFrame,
    pod_files: Dict[str, Path],
    present: Set[str],
    missing: Set[str],
    extra: Set[str]
) -> pd.DataFrame:
    """
    Create report DataFrame with all POD statuses.

    Args:
        manifest_df: Original manifest DataFrame
        pod_files: Dict of delivery_id to file path
        present: Set of present delivery IDs
        missing: Set of missing delivery IDs
        extra: Set of extra delivery IDs

    Returns:
        Report DataFrame
    """
    rows = []

    # Process manifest entries
    for _, row in manifest_df.iterrows():
        delivery_id = str(row.get('delivery_id', '')).strip()

        if delivery_id in present:
            status = 'Present'
            filename = pod_files.get(delivery_id, Path()).name
        elif delivery_id in missing:
            status = 'Missing'
            filename = ''
        else:
            continue

        rows.append({
            'Delivery ID': delivery_id,
            'Status': status,
            'Filename': filename,
            'Manifest Date': row.get('date', ''),
            'Customer': row.get('customer', ''),
        })

    # Add extra PODs (not in manifest)
    for delivery_id in extra:
        file_path = pod_files.get(delivery_id, Path())
        rows.append({
            'Delivery ID': delivery_id,
            'Status': 'Extra',
            'Filename': file_path.name if file_path else '',
            'Manifest Date': 'N/A',
            'Customer': 'N/A (not in manifest)',
        })

    df = pd.DataFrame(rows)

    # Sort: Missing first, then Present, then Extra
    status_order = {'Missing': 0, 'Present': 1, 'Extra': 2}
    df['_sort'] = df['Status'].map(status_order)
    df = df.sort_values(['_sort', 'Delivery ID']).drop('_sort', axis=1)

    return df


def run_pod_check(
    pod_folder: str = None,
    manifest_path: str = None,
    output_folder: str = None
) -> Path:
    """
    Main function to run POD presence check.

    Args:
        pod_folder: Path to folder with POD PDFs
        manifest_path: Path to manifest Excel
        output_folder: Path for output report

    Returns:
        Path to generated report
    """
    # Initialize config
    config = PODConfig(
        pod_folder=pod_folder,
        manifest_path=manifest_path,
        output_folder=output_folder
    )

    # Validate paths
    issues = config.validate_paths()
    if issues:
        for key, msg in issues.items():
            print(f"Error: {msg}")
        sys.exit(1)

    print(f"POD Folder: {config.pod_folder}")
    print(f"Manifest: {config.manifest_path}")
    print("-" * 50)

    # Read manifest
    print("Reading manifest...")
    manifest_df = read_manifest(config.manifest_path, PODConfig.MANIFEST_COLUMNS)
    manifest_ids = set(manifest_df['delivery_id'].astype(str).str.strip())
    print(f"Found {len(manifest_ids)} entries in manifest")

    # Scan POD folder
    print("Scanning POD folder...")
    pod_files = scan_pod_folder(config.pod_folder, PODConfig.POD_FILE_EXTENSIONS)
    scanned_ids = set(pod_files.keys())
    print(f"Found {len(scanned_ids)} POD files")

    # Compare
    print("Comparing...")
    present, missing, extra = compare_pods(manifest_ids, scanned_ids)

    # Create report
    report_df = create_report_dataframe(
        manifest_df, pod_files, present, missing, extra
    )

    # Summary statistics
    summary = {
        'Total in Manifest': len(manifest_ids),
        'PODs Present': len(present),
        'PODs Missing': len(missing),
        'Extra PODs': len(extra),
        'Match Rate': f"{len(present) / len(manifest_ids) * 100:.1f}%" if manifest_ids else "N/A",
        'Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

    # Print summary
    print("-" * 50)
    print("SUMMARY")
    print("-" * 50)
    for key, value in summary.items():
        print(f"{key}: {value}")

    # Write report
    output_path = config.get_output_path('pod_check_report')
    write_report(report_df, output_path, 'POD Check', summary)

    # Apply status formatting
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    ws = wb.active
    # Find the Status column and apply formatting
    status_col = None
    for idx, cell in enumerate(ws[8], 1):  # Row 8 is typically header after summary
        if cell.value == 'Status':
            status_col = idx
            break
    if status_col:
        apply_status_formatting(ws, status_col, 9)
    wb.save(output_path)

    print("-" * 50)
    print(f"Report saved: {output_path}")

    return output_path


def main():
    """Parse arguments and run POD check."""
    parser = argparse.ArgumentParser(
        description='POD Presence Validator - Compare scanned PODs against manifest'
    )
    parser.add_argument(
        'pod_folder',
        nargs='?',
        help='Path to folder containing POD PDFs'
    )
    parser.add_argument(
        'manifest_path',
        nargs='?',
        help='Path to manifest Excel file'
    )
    parser.add_argument(
        '--output', '-o',
        help='Output folder for report'
    )

    args = parser.parse_args()

    run_pod_check(
        pod_folder=args.pod_folder,
        manifest_path=args.manifest_path,
        output_folder=args.output
    )


if __name__ == '__main__':
    main()
