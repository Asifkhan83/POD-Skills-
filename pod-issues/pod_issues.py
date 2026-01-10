"""
POD Quality Checker - /pod-issues
Detect common POD issues (date mismatch, stamp problems, customer mismatch).

Usage:
    python pod_issues.py [pod_folder] [manifest_path] [--ocr]
"""
import sys
import re
import argparse
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple

import pandas as pd

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))
from shared.config import PODConfig, parse_delivery_id
from shared.excel_utils import write_report

# Try importing PDF libraries
try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False
    print("Warning: pdfplumber not installed. Install with: pip install pdfplumber")

try:
    from fuzzywuzzy import fuzz
    HAS_FUZZY = True
except ImportError:
    HAS_FUZZY = False
    print("Warning: fuzzywuzzy not installed. Install with: pip install fuzzywuzzy")


# Date patterns to look for in PDFs
DATE_PATTERNS = [
    r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',  # DD/MM/YYYY, MM-DD-YY
    r'\d{4}[/-]\d{1,2}[/-]\d{1,2}',     # YYYY-MM-DD
    r'\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\s+\d{2,4}',  # 15 January 2024
]


def extract_pdf_text(pdf_path: Path) -> str:
    """
    Extract text content from PDF.

    Args:
        pdf_path: Path to PDF file

    Returns:
        Extracted text content
    """
    if not HAS_PDFPLUMBER:
        return ""

    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
            return text
    except Exception as e:
        print(f"Warning: Could not read {pdf_path.name}: {e}")
        return ""


def extract_dates_from_text(text: str) -> List[str]:
    """
    Extract date strings from text.

    Args:
        text: Text to search

    Returns:
        List of found date strings
    """
    dates = []
    for pattern in DATE_PATTERNS:
        matches = re.findall(pattern, text, re.IGNORECASE)
        dates.extend(matches)
    return dates


def parse_date(date_str: str) -> Optional[datetime]:
    """
    Parse date string to datetime object.

    Args:
        date_str: Date string in various formats

    Returns:
        datetime object or None if parsing fails
    """
    formats = [
        '%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d', '%d-%m-%Y',
        '%d/%m/%y', '%m/%d/%y', '%Y/%m/%d',
        '%d %B %Y', '%d %b %Y', '%B %d, %Y'
    ]

    for fmt in formats:
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except ValueError:
            continue

    return None


def check_date_issue(
    pdf_text: str,
    manifest_date: datetime,
    tolerance_days: int = 2
) -> Optional[Dict]:
    """
    Check for date mismatch between PDF and manifest.

    Args:
        pdf_text: Extracted PDF text
        manifest_date: Expected date from manifest
        tolerance_days: Allowed difference in days

    Returns:
        Issue dict if found, None otherwise
    """
    if not manifest_date:
        return None

    pdf_dates = extract_dates_from_text(pdf_text)

    if not pdf_dates:
        return {
            'type': 'Date Issue',
            'severity': 'Medium',
            'details': 'No date found in PDF',
            'manifest_value': manifest_date.strftime('%Y-%m-%d') if manifest_date else '',
            'pdf_value': 'Not found'
        }

    # Check if any PDF date matches manifest date
    for date_str in pdf_dates:
        pdf_date = parse_date(date_str)
        if pdf_date:
            diff = abs((pdf_date - manifest_date).days)
            if diff <= tolerance_days:
                return None  # Date matches within tolerance

            return {
                'type': 'Date Mismatch',
                'severity': 'High' if diff > 7 else 'Medium',
                'details': f'Date differs by {diff} days',
                'manifest_value': manifest_date.strftime('%Y-%m-%d'),
                'pdf_value': date_str
            }

    return None


def check_customer_match(
    pdf_text: str,
    manifest_customer: str,
    threshold: int = 80
) -> Optional[Dict]:
    """
    Check if customer name in PDF matches manifest.

    Args:
        pdf_text: Extracted PDF text
        manifest_customer: Expected customer name
        threshold: Fuzzy match threshold (0-100)

    Returns:
        Issue dict if mismatch found, None otherwise
    """
    if not manifest_customer or not HAS_FUZZY:
        return None

    manifest_customer = manifest_customer.strip().upper()

    # Simple heuristic: look for customer name in text
    pdf_upper = pdf_text.upper()

    # Check for exact match first
    if manifest_customer in pdf_upper:
        return None

    # Fuzzy match against text segments
    words = pdf_text.split()
    best_match = 0

    # Check 2-4 word combinations
    for i in range(len(words)):
        for j in range(i + 1, min(i + 5, len(words) + 1)):
            segment = ' '.join(words[i:j]).upper()
            ratio = fuzz.ratio(manifest_customer, segment)
            best_match = max(best_match, ratio)

    if best_match >= threshold:
        return None

    return {
        'type': 'Customer Mismatch',
        'severity': 'High' if best_match < 50 else 'Medium',
        'details': f'Best match: {best_match}%',
        'manifest_value': manifest_customer,
        'pdf_value': f'No match found (best: {best_match}%)'
    }


def check_stamp_presence(pdf_path: Path) -> Optional[Dict]:
    """
    Basic check for stamp/signature presence.
    Uses heuristics based on PDF content.

    Args:
        pdf_path: Path to PDF file

    Returns:
        Issue dict if stamp appears missing, None otherwise
    """
    if not HAS_PDFPLUMBER:
        return None

    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Check for images (stamps are often images)
            has_images = False
            for page in pdf.pages:
                if page.images:
                    has_images = True
                    break

            if not has_images:
                return {
                    'type': 'Stamp Check',
                    'severity': 'Low',
                    'details': 'No images/stamps detected in PDF',
                    'manifest_value': 'Expected stamp',
                    'pdf_value': 'No images found'
                }

    except Exception as e:
        return {
            'type': 'Stamp Check',
            'severity': 'Low',
            'details': f'Could not analyze: {e}',
            'manifest_value': '',
            'pdf_value': 'Analysis failed'
        }

    return None


def analyze_pod(
    pdf_path: Path,
    manifest_row: Dict,
    config: PODConfig
) -> List[Dict]:
    """
    Analyze a single POD for issues.

    Args:
        pdf_path: Path to POD PDF
        manifest_row: Manifest data for this delivery
        config: Configuration object

    Returns:
        List of issues found
    """
    issues = []
    delivery_id = parse_delivery_id(pdf_path.name)

    # Extract PDF text
    pdf_text = extract_pdf_text(pdf_path)

    # Check date
    manifest_date = manifest_row.get('date')
    if isinstance(manifest_date, str):
        manifest_date = parse_date(manifest_date)
    elif isinstance(manifest_date, datetime):
        pass
    else:
        manifest_date = None

    date_issue = check_date_issue(
        pdf_text, manifest_date, config.DATE_TOLERANCE_DAYS
    )
    if date_issue:
        date_issue['delivery_id'] = delivery_id
        issues.append(date_issue)

    # Check customer
    manifest_customer = manifest_row.get('customer', '')
    customer_issue = check_customer_match(
        pdf_text, manifest_customer, config.CUSTOMER_MATCH_THRESHOLD
    )
    if customer_issue:
        customer_issue['delivery_id'] = delivery_id
        issues.append(customer_issue)

    # Check stamp
    stamp_issue = check_stamp_presence(pdf_path)
    if stamp_issue:
        stamp_issue['delivery_id'] = delivery_id
        issues.append(stamp_issue)

    return issues


def run_pod_issues(
    pod_folder: str = None,
    manifest_path: str = None,
    output_folder: str = None,
    use_ocr: bool = False
) -> Path:
    """
    Main function to run POD issue detection.

    Args:
        pod_folder: Path to folder with POD PDFs
        manifest_path: Path to manifest Excel
        output_folder: Path for output report
        use_ocr: Whether to use OCR for scanned PDFs

    Returns:
        Path to generated report
    """
    # Check dependencies
    if not HAS_PDFPLUMBER:
        print("Error: pdfplumber is required. Install with: pip install pdfplumber")
        sys.exit(1)

    # Initialize config
    config = PODConfig(
        pod_folder=pod_folder,
        manifest_path=manifest_path,
        output_folder=output_folder
    )

    # Validate paths
    issues = config.validate_paths()
    if issues:
        for key, msg in issues.items():
            print(f"Error: {msg}")
        sys.exit(1)

    print(f"POD Folder: {config.pod_folder}")
    print(f"Manifest: {config.manifest_path}")
    print("-" * 50)

    # Read manifest
    print("Reading manifest...")
    manifest_df = pd.read_excel(config.manifest_path)

    # Standardize columns
    column_map = {v: k for k, v in PODConfig.MANIFEST_COLUMNS.items() if v in manifest_df.columns}
    manifest_df = manifest_df.rename(columns=column_map)

    if 'delivery_id' in manifest_df.columns:
        manifest_df['delivery_id'] = manifest_df['delivery_id'].astype(str).str.strip()

    # Create lookup dict
    manifest_lookup = {
        str(row['delivery_id']): row.to_dict()
        for _, row in manifest_df.iterrows()
        if 'delivery_id' in row
    }

    print(f"Found {len(manifest_lookup)} entries in manifest")

    # Scan POD folder
    print("Analyzing POD files...")
    all_issues = []
    pdf_count = 0

    for ext in config.POD_FILE_EXTENSIONS:
        for pdf_path in config.pod_folder.glob(f"*{ext}"):
            pdf_count += 1
            delivery_id = parse_delivery_id(pdf_path.name)

            if delivery_id in manifest_lookup:
                manifest_row = manifest_lookup[delivery_id]
                pod_issues = analyze_pod(pdf_path, manifest_row, config)
                all_issues.extend(pod_issues)

            if pdf_count % 50 == 0:
                print(f"  Processed {pdf_count} files...")

    print(f"Analyzed {pdf_count} POD files")
    print(f"Found {len(all_issues)} issues")

    # Create report DataFrame
    if all_issues:
        report_df = pd.DataFrame(all_issues)
        report_df = report_df.rename(columns={
            'delivery_id': 'Delivery ID',
            'type': 'Issue Type',
            'severity': 'Severity',
            'details': 'Details',
            'manifest_value': 'Expected Value',
            'pdf_value': 'PDF Value'
        })

        # Add needs action column
        report_df['Needs Action'] = report_df['Severity'].apply(
            lambda x: 'Yes' if x in ['High', 'Medium'] else 'No'
        )

        # Sort by severity
        severity_order = {'High': 0, 'Medium': 1, 'Low': 2}
        report_df['_sort'] = report_df['Severity'].map(severity_order)
        report_df = report_df.sort_values(['_sort', 'Delivery ID']).drop('_sort', axis=1)
    else:
        report_df = pd.DataFrame(columns=[
            'Delivery ID', 'Issue Type', 'Severity', 'Details',
            'Expected Value', 'PDF Value', 'Needs Action'
        ])

    # Calculate summary
    issue_counts = report_df['Issue Type'].value_counts().to_dict() if len(report_df) > 0 else {}
    severity_counts = report_df['Severity'].value_counts().to_dict() if len(report_df) > 0 else {}

    summary = {
        'PODs Analyzed': pdf_count,
        'Total Issues': len(all_issues),
        'High Severity': severity_counts.get('High', 0),
        'Medium Severity': severity_counts.get('Medium', 0),
        'Low Severity': severity_counts.get('Low', 0),
        **{f'{k} Issues': v for k, v in issue_counts.items()},
        'Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

    # Print summary
    print("-" * 50)
    print("SUMMARY")
    print("-" * 50)
    for key, value in summary.items():
        print(f"{key}: {value}")

    # Write report
    output_path = config.get_output_path('pod_issues_report')
    write_report(report_df, output_path, 'POD Issues', summary)

    # Apply severity formatting
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    wb = load_workbook(output_path)
    ws = wb.active

    high_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    low_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Find severity column and apply formatting
    for row in ws.iter_rows(min_row=len(summary) + 3, max_row=ws.max_row):
        severity_cell = row[2]  # Severity is 3rd column
        severity = str(severity_cell.value).strip() if severity_cell.value else ''

        fill = None
        if severity == 'High':
            fill = high_fill
        elif severity == 'Medium':
            fill = medium_fill
        elif severity == 'Low':
            fill = low_fill

        if fill:
            for cell in row:
                cell.fill = fill

    wb.save(output_path)

    print("-" * 50)
    print(f"Report saved: {output_path}")

    return output_path


def main():
    """Parse arguments and run POD issues detection."""
    parser = argparse.ArgumentParser(
        description='POD Quality Checker - Detect issues in POD PDFs'
    )
    parser.add_argument(
        'pod_folder',
        nargs='?',
        help='Path to folder containing POD PDFs'
    )
    parser.add_argument(
        'manifest_path',
        nargs='?',
        help='Path to manifest Excel file'
    )
    parser.add_argument(
        '--output', '-o',
        help='Output folder for report'
    )
    parser.add_argument(
        '--ocr',
        action='store_true',
        help='Enable OCR for scanned PDFs'
    )

    args = parser.parse_args()

    run_pod_issues(
        pod_folder=args.pod_folder,
        manifest_path=args.manifest_path,
        output_folder=args.output,
        use_ocr=args.ocr
    )


if __name__ == '__main__':
    main()
