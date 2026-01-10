"""
POD Archive Manager - /pod-archive
Organize and archive processed PODs by date, customer, or status.

Usage:
    python pod_archive.py [source_folder] [archive_folder] [--mode mode]
"""
import sys
import shutil
import argparse
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))
from shared.config import PODConfig, parse_delivery_id
from shared.excel_utils import write_report


def get_file_date(file_path: Path) -> datetime:
    """
    Get file modification date.

    Args:
        file_path: Path to file

    Returns:
        datetime of file modification
    """
    return datetime.fromtimestamp(file_path.stat().st_mtime)


def determine_archive_path_by_date(
    file_path: Path,
    archive_root: Path,
    manifest_data: Dict = None
) -> Path:
    """
    Determine archive path based on date.

    Args:
        file_path: Source file path
        archive_root: Archive root folder
        manifest_data: Optional manifest data with date

    Returns:
        Destination path
    """
    # Try to get date from manifest first
    delivery_id = parse_delivery_id(file_path.name)
    date = None

    if manifest_data and delivery_id in manifest_data:
        date_val = manifest_data[delivery_id].get('date')
        if isinstance(date_val, datetime):
            date = date_val
        elif isinstance(date_val, str):
            try:
                date = datetime.strptime(date_val, '%Y-%m-%d')
            except ValueError:
                pass

    # Fall back to file date
    if not date:
        date = get_file_date(file_path)

    # Create path: Archive/YYYY/MM/DD/
    dest_folder = archive_root / str(date.year) / f"{date.month:02d}" / f"{date.day:02d}"
    return dest_folder / file_path.name


def determine_archive_path_by_customer(
    file_path: Path,
    archive_root: Path,
    manifest_data: Dict
) -> Path:
    """
    Determine archive path based on customer.

    Args:
        file_path: Source file path
        archive_root: Archive root folder
        manifest_data: Manifest data with customer info

    Returns:
        Destination path
    """
    delivery_id = parse_delivery_id(file_path.name)
    customer = "Unknown"
    year_month = datetime.now().strftime('%Y-%m')

    if manifest_data and delivery_id in manifest_data:
        customer_val = manifest_data[delivery_id].get('customer', '')
        if customer_val:
            # Clean customer name for folder
            customer = str(customer_val).strip()
            customer = "".join(c for c in customer if c.isalnum() or c in ' -_')
            customer = customer[:50]  # Limit length

        date_val = manifest_data[delivery_id].get('date')
        if isinstance(date_val, datetime):
            year_month = date_val.strftime('%Y-%m')

    # Create path: Archive/CustomerName/YYYY-MM/
    dest_folder = archive_root / customer / year_month
    return dest_folder / file_path.name


def determine_archive_path_by_status(
    file_path: Path,
    archive_root: Path,
    status_data: Dict
) -> Path:
    """
    Determine archive path based on status.

    Args:
        file_path: Source file path
        archive_root: Archive root folder
        status_data: Status data from status report

    Returns:
        Destination path
    """
    delivery_id = parse_delivery_id(file_path.name)
    status = "Unknown"

    if status_data and delivery_id in status_data:
        resolution = status_data[delivery_id].get('resolution_status', '')

        if resolution in ['Closed', 'Ready to Close']:
            status = "Completed"
        elif resolution == 'Has Issues':
            status = "Issues"
        elif resolution == 'Pending POD':
            status = "Pending"
        else:
            status = resolution or "Unknown"

    # Create path: Archive/Status/YYYY-MM/
    year_month = datetime.now().strftime('%Y-%m')
    dest_folder = archive_root / status / year_month
    return dest_folder / file_path.name


def load_manifest_data(manifest_path: Path) -> Dict:
    """
    Load manifest data into lookup dict.

    Args:
        manifest_path: Path to manifest Excel

    Returns:
        Dict mapping delivery_id to row data
    """
    if not manifest_path or not manifest_path.exists():
        return {}

    df = pd.read_excel(manifest_path)

    # Standardize columns
    column_map = {v: k for k, v in PODConfig.MANIFEST_COLUMNS.items() if v in df.columns}
    df = df.rename(columns=column_map)

    if 'delivery_id' not in df.columns:
        return {}

    df['delivery_id'] = df['delivery_id'].astype(str).str.strip()

    return {
        row['delivery_id']: row.to_dict()
        for _, row in df.iterrows()
    }


def load_status_data(status_path: Path) -> Dict:
    """
    Load status report data into lookup dict.

    Args:
        status_path: Path to status report Excel

    Returns:
        Dict mapping delivery_id to status data
    """
    if not status_path or not status_path.exists():
        return {}

    # Find the header row dynamically
    df_raw = pd.read_excel(status_path, header=None)
    header_row = 0
    for i, row in df_raw.iterrows():
        if 'Delivery ID' in str(row.values):
            header_row = i
            break

    df = pd.read_excel(status_path, skiprows=header_row)

    status_data = {}
    for _, row in df.iterrows():
        delivery_id = str(row.get('Delivery ID', '')).strip()
        if delivery_id and delivery_id != 'nan':
            status_data[delivery_id] = {
                'pod_received': row.get('POD Received', ''),
                'has_issues': row.get('Has Issues', ''),
                'resolution_status': row.get('Resolution Status', ''),
                'ready_to_close': row.get('Ready to Close', '')
            }

    return status_data


def archive_files(
    source_folder: Path,
    archive_folder: Path,
    mode: str = 'by-date',
    manifest_path: Path = None,
    status_path: Path = None,
    copy_files: bool = False,
    dry_run: bool = False
) -> Tuple[List[Dict], List[Dict]]:
    """
    Archive POD files based on specified mode.

    Args:
        source_folder: Source folder with PODs
        archive_folder: Destination archive folder
        mode: Archive mode (by-date, by-customer, by-status)
        manifest_path: Path to manifest Excel
        status_path: Path to status report
        copy_files: Copy instead of move
        dry_run: Preview only, don't move

    Returns:
        Tuple of (success_list, error_list)
    """
    # Load data based on mode
    manifest_data = {}
    status_data = {}

    if mode in ['by-date', 'by-customer']:
        manifest_data = load_manifest_data(manifest_path)

    if mode == 'by-status':
        status_data = load_status_data(status_path)

    success_list = []
    error_list = []

    # Process files
    for ext in PODConfig.POD_FILE_EXTENSIONS:
        for file_path in source_folder.glob(f"*{ext}"):
            try:
                # Determine destination
                if mode == 'by-date':
                    dest_path = determine_archive_path_by_date(
                        file_path, archive_folder, manifest_data
                    )
                elif mode == 'by-customer':
                    dest_path = determine_archive_path_by_customer(
                        file_path, archive_folder, manifest_data
                    )
                elif mode == 'by-status':
                    dest_path = determine_archive_path_by_status(
                        file_path, archive_folder, status_data
                    )
                else:
                    dest_path = archive_folder / file_path.name

                if dry_run:
                    action = 'Would copy' if copy_files else 'Would move'
                else:
                    # Create destination folder
                    dest_path.parent.mkdir(parents=True, exist_ok=True)

                    # Copy or move
                    if copy_files:
                        shutil.copy2(file_path, dest_path)
                        action = 'Copied'
                    else:
                        shutil.move(str(file_path), str(dest_path))
                        action = 'Moved'

                success_list.append({
                    'Delivery ID': parse_delivery_id(file_path.name),
                    'Source': str(file_path),
                    'Destination': str(dest_path),
                    'Action': action,
                    'Archive Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })

            except Exception as e:
                error_list.append({
                    'Delivery ID': parse_delivery_id(file_path.name),
                    'Source': str(file_path),
                    'Error': str(e)
                })

    return success_list, error_list


def run_pod_archive(
    source_folder: str,
    archive_folder: str,
    mode: str = 'by-date',
    manifest_path: str = None,
    status_path: str = None,
    copy_files: bool = False,
    dry_run: bool = False,
    output_folder: str = None
) -> Path:
    """
    Main function to run POD archiving.

    Args:
        source_folder: Path to source folder
        archive_folder: Path to archive folder
        mode: Archive mode
        manifest_path: Path to manifest Excel
        status_path: Path to status report
        copy_files: Copy instead of move
        dry_run: Preview only
        output_folder: Path for log output

    Returns:
        Path to archive log
    """
    source = Path(source_folder)
    archive = Path(archive_folder)

    if not source.exists():
        print(f"Error: Source folder not found: {source}")
        sys.exit(1)

    # Validate mode requirements
    if mode == 'by-customer' and not manifest_path:
        print("Warning: by-customer mode requires manifest for best results")

    if mode == 'by-status' and not status_path:
        print("Warning: by-status mode requires status report for best results")

    print(f"Source: {source}")
    print(f"Archive: {archive}")
    print(f"Mode: {mode}")
    print(f"Action: {'Copy' if copy_files else 'Move'}")
    print(f"Dry Run: {dry_run}")
    print("-" * 50)

    # Run archiving
    print("Processing files...")
    success_list, error_list = archive_files(
        source_folder=source,
        archive_folder=archive,
        mode=mode,
        manifest_path=Path(manifest_path) if manifest_path else None,
        status_path=Path(status_path) if status_path else None,
        copy_files=copy_files,
        dry_run=dry_run
    )

    print(f"Processed: {len(success_list)} files")
    print(f"Errors: {len(error_list)} files")

    # Create summary
    summary = {
        'Files Processed': len(success_list),
        'Errors': len(error_list),
        'Mode': mode,
        'Action': 'Copy' if copy_files else 'Move',
        'Dry Run': 'Yes' if dry_run else 'No',
        'Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

    # Print summary
    print("-" * 50)
    print("SUMMARY")
    print("-" * 50)
    for key, value in summary.items():
        print(f"{key}: {value}")

    # Create log DataFrame
    if success_list:
        log_df = pd.DataFrame(success_list)
    else:
        log_df = pd.DataFrame(columns=[
            'Delivery ID', 'Source', 'Destination', 'Action', 'Archive Date'
        ])

    # Write log
    config = PODConfig(output_folder=output_folder)
    output_path = config.get_output_path('pod_archive_log')
    write_report(log_df, output_path, 'Archive Log', summary)

    # Add error sheet if any
    if error_list:
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        error_df = pd.DataFrame(error_list)

        ws_errors = wb.create_sheet('Errors')
        from openpyxl.utils.dataframe import dataframe_to_rows
        for r_idx, row in enumerate(dataframe_to_rows(error_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_errors.cell(row=r_idx, column=c_idx, value=value)

        wb.save(output_path)

    print("-" * 50)
    print(f"Archive log saved: {output_path}")

    return output_path


def main():
    """Parse arguments and run POD archiving."""
    parser = argparse.ArgumentParser(
        description='POD Archive Manager - Organize and archive PODs'
    )
    parser.add_argument(
        'source_folder',
        help='Path to folder with PODs to archive'
    )
    parser.add_argument(
        'archive_folder',
        help='Destination archive folder'
    )
    parser.add_argument(
        '--mode', '-m',
        choices=['by-date', 'by-customer', 'by-status'],
        default='by-date',
        help='Archive organization mode'
    )
    parser.add_argument(
        '--manifest',
        help='Path to manifest Excel'
    )
    parser.add_argument(
        '--status-report',
        help='Path to status report'
    )
    parser.add_argument(
        '--copy',
        action='store_true',
        help='Copy files instead of moving'
    )
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Preview without moving files'
    )
    parser.add_argument(
        '--output', '-o',
        help='Output folder for archive log'
    )

    args = parser.parse_args()

    run_pod_archive(
        source_folder=args.source_folder,
        archive_folder=args.archive_folder,
        mode=args.mode,
        manifest_path=args.manifest,
        status_path=args.status_report,
        copy_files=args.copy,
        dry_run=args.dry_run,
        output_folder=args.output
    )


if __name__ == '__main__':
    main()
