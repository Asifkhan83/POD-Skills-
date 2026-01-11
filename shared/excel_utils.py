"""
Excel utilities for POD skills.
Common operations for reading, writing, and formatting Excel files.
"""
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# Style definitions
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PRESENT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MISSING_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
EXTRA_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ISSUE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def find_header_row(filepath: Path, marker_column: str = 'Delivery ID') -> int:
    """
    Find the header row in an Excel file with summary sections.

    Args:
        filepath: Path to Excel file
        marker_column: Column name to look for in header

    Returns:
        Row number where header is found (0-indexed), or 0 if not found
    """
    df_raw = pd.read_excel(filepath, header=None)

    for i, row in df_raw.iterrows():
        row_values = [str(v) for v in row.values if pd.notna(v)]
        if marker_column in row_values:
            return i

    return 0


def read_report_with_summary(filepath: Path, marker_column: str = 'Delivery ID') -> pd.DataFrame:
    """
    Read an Excel report that has a summary section before the data.

    Args:
        filepath: Path to Excel file
        marker_column: Column name to identify the header row

    Returns:
        DataFrame with the data section only
    """
    header_row = find_header_row(filepath, marker_column)
    return pd.read_excel(filepath, skiprows=header_row)


def validate_manifest_columns(df: pd.DataFrame, required_columns: List[str]) -> Dict[str, str]:
    """
    Validate that required columns exist in DataFrame.

    Args:
        df: DataFrame to validate
        required_columns: List of required column names

    Returns:
        Dict of validation issues (empty if all valid)
    """
    issues = {}
    missing = [col for col in required_columns if col not in df.columns]

    if missing:
        issues['missing_columns'] = f"Missing required columns: {', '.join(missing)}"

    # Check for empty delivery_id column
    if 'delivery_id' in df.columns:
        blank_count = df['delivery_id'].isna().sum() + (df['delivery_id'] == '').sum()
        if blank_count > 0:
            issues['blank_ids'] = f"Found {blank_count} blank delivery ID(s)"

        # Check for duplicates
        duplicates = df['delivery_id'].duplicated().sum()
        if duplicates > 0:
            issues['duplicate_ids'] = f"Found {duplicates} duplicate delivery ID(s) in manifest"

    return issues


def read_manifest(filepath: Path, columns: Dict[str, str], validate: bool = True) -> pd.DataFrame:
    """
    Read manifest Excel file and standardize column names.

    Args:
        filepath: Path to manifest Excel file
        columns: Mapping of standard names to actual column names
        validate: Whether to validate and warn about issues

    Returns:
        DataFrame with standardized columns
    """
    df = pd.read_excel(filepath)

    # Create reverse mapping and rename columns
    rename_map = {v: k for k, v in columns.items() if v in df.columns}
    df = df.rename(columns=rename_map)

    # Ensure delivery_id is string
    if 'delivery_id' in df.columns:
        df['delivery_id'] = df['delivery_id'].astype(str).str.strip()

    # Validate if requested
    if validate:
        issues = validate_manifest_columns(df, ['delivery_id'])
        for issue_type, message in issues.items():
            print(f"Warning: {message}")

    return df


def write_report(
    df: pd.DataFrame,
    filepath: Path,
    sheet_name: str = 'Report',
    summary: Optional[Dict[str, Any]] = None
) -> Path:
    """
    Write DataFrame to Excel with formatting.

    Args:
        df: DataFrame to write
        filepath: Output file path
        sheet_name: Name of the worksheet
        summary: Optional summary dict to add at top

    Returns:
        Path to created file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    start_row = 1

    # Add summary section if provided
    if summary:
        ws['A1'] = 'Summary'
        ws['A1'].font = Font(bold=True, size=14)
        start_row = 2

        for key, value in summary.items():
            ws.cell(row=start_row, column=1, value=key)
            ws.cell(row=start_row, column=2, value=value)
            start_row += 1

        start_row += 1  # Empty row before data

    # Write DataFrame
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER

            # Header formatting
            if r_idx == start_row:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center')

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filepath)
    return filepath


def apply_status_formatting(ws, status_column: int, start_row: int):
    """
    Apply conditional formatting based on status values.

    Args:
        ws: Worksheet object
        status_column: Column number containing status
        start_row: First data row (after header)
    """
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        status_cell = row[status_column - 1]
        status = str(status_cell.value).lower() if status_cell.value else ''

        fill = None
        if status == 'present' or status == 'received':
            fill = PRESENT_FILL
        elif status == 'missing':
            fill = MISSING_FILL
        elif status == 'extra':
            fill = EXTRA_FILL
        elif 'issue' in status or 'error' in status:
            fill = ISSUE_FILL

        if fill:
            for cell in row:
                cell.fill = fill


def create_summary_dict(df: pd.DataFrame, status_column: str = 'status') -> Dict[str, int]:
    """
    Create summary statistics from DataFrame.

    Args:
        df: DataFrame with status column
        status_column: Name of status column

    Returns:
        Dict with counts per status
    """
    summary = {'Total': len(df)}

    if status_column in df.columns:
        status_counts = df[status_column].value_counts().to_dict()
        summary.update(status_counts)

    summary['Generated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    return summary


def merge_reports(reports: List[Path], output_path: Path) -> Path:
    """
    Merge multiple Excel reports into one workbook.

    Args:
        reports: List of report file paths
        output_path: Output file path

    Returns:
        Path to merged file
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for report_path in reports:
        df = pd.read_excel(report_path)
        sheet_name = report_path.stem[:31]  # Excel sheet name limit

        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(output_path)
    return output_path
