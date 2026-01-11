"""
Report utilities for POD skills.
Generates Markdown reports with export options for CSV, PDF, HTML, Excel.
"""
import csv
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
import pandas as pd


class MarkdownReport:
    """Generate Markdown reports with export capabilities."""

    def __init__(self, title: str, generated_at: datetime = None):
        self.title = title
        self.generated_at = generated_at or datetime.now()
        self.sections: List[Dict[str, Any]] = []
        self.summary: Dict[str, Any] = {}
        self.dataframe: Optional[pd.DataFrame] = None

    def set_summary(self, summary: Dict[str, Any]):
        """Set the summary statistics."""
        self.summary = summary

    def set_data(self, df: pd.DataFrame):
        """Set the main data table."""
        self.dataframe = df

    def add_section(self, title: str, content: str):
        """Add a custom section."""
        self.sections.append({'title': title, 'content': content})

    def _format_value(self, value) -> str:
        """Format a value for display."""
        if pd.isna(value) or value is None:
            return '-'
        if isinstance(value, float):
            return f"{value:.1f}" if value != int(value) else str(int(value))
        return str(value)

    def _dataframe_to_markdown(self, df: pd.DataFrame) -> str:
        """Convert DataFrame to Markdown table."""
        if df is None or df.empty:
            return "*No data available*\n"

        lines = []

        # Header
        headers = list(df.columns)
        lines.append('| ' + ' | '.join(headers) + ' |')
        lines.append('|' + '|'.join(['---' for _ in headers]) + '|')

        # Rows
        for _, row in df.iterrows():
            values = [self._format_value(row[col]) for col in headers]
            lines.append('| ' + ' | '.join(values) + ' |')

        return '\n'.join(lines)

    def to_markdown(self) -> str:
        """Generate full Markdown report."""
        lines = []

        # Title
        lines.append(f"# {self.title}")
        lines.append("")
        lines.append(f"**Generated:** {self.generated_at.strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")

        # Summary
        if self.summary:
            lines.append("## Summary")
            lines.append("")
            lines.append("| Metric | Value |")
            lines.append("|--------|-------|")
            for key, value in self.summary.items():
                if key != 'Generated':
                    lines.append(f"| {key} | {self._format_value(value)} |")
            lines.append("")

        # Custom sections
        for section in self.sections:
            lines.append(f"## {section['title']}")
            lines.append("")
            lines.append(section['content'])
            lines.append("")

        # Main data table
        if self.dataframe is not None and not self.dataframe.empty:
            lines.append("## Details")
            lines.append("")
            lines.append(self._dataframe_to_markdown(self.dataframe))
            lines.append("")

        return '\n'.join(lines)

    def save_markdown(self, filepath: Path) -> Path:
        """Save report as Markdown file."""
        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(self.to_markdown())

        return filepath

    def save_csv(self, filepath: Path) -> Path:
        """Export data to CSV file."""
        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)

        if self.dataframe is not None:
            self.dataframe.to_csv(filepath, index=False, encoding='utf-8')

        return filepath

    def save_excel(self, filepath: Path) -> Path:
        """Export to Excel with formatting."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = self.title[:31]  # Excel sheet name limit

        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        row_num = 1

        # Title
        ws.cell(row=row_num, column=1, value=self.title)
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
        row_num += 1

        # Summary
        if self.summary:
            row_num += 1
            ws.cell(row=row_num, column=1, value='Summary')
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
            row_num += 1

            for key, value in self.summary.items():
                ws.cell(row=row_num, column=1, value=key)
                ws.cell(row=row_num, column=2, value=value)
                row_num += 1

        # Data table
        if self.dataframe is not None and not self.dataframe.empty:
            row_num += 1
            start_row = row_num

            for r_idx, row in enumerate(dataframe_to_rows(self.dataframe, index=False, header=True), start_row):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = thin_border

                    if r_idx == start_row:  # Header row
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')
                row_num = r_idx

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        wb.save(filepath)
        return filepath

    def save_html(self, filepath: Path) -> Path:
        """Export to HTML file."""
        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)

        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{self.title}</title>
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 40px; background: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        h1 {{ color: #333; border-bottom: 2px solid #366092; padding-bottom: 10px; }}
        h2 {{ color: #366092; margin-top: 30px; }}
        .generated {{ color: #666; font-size: 14px; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
        th {{ background: #366092; color: white; padding: 12px; text-align: left; }}
        td {{ padding: 10px; border-bottom: 1px solid #ddd; }}
        tr:hover {{ background: #f9f9f9; }}
        .summary-table {{ width: auto; }}
        .summary-table td {{ padding: 8px 20px 8px 0; }}
        .status-present {{ background: #c6efce; }}
        .status-missing {{ background: #ffc7ce; }}
        .status-extra {{ background: #ffeb9c; }}
        .match-yes {{ color: #006600; font-weight: bold; }}
        .match-no {{ color: #cc0000; }}
        .match-partial {{ color: #cc6600; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{self.title}</h1>
        <p class="generated">Generated: {self.generated_at.strftime('%Y-%m-%d %H:%M:%S')}</p>
"""

        # Summary
        if self.summary:
            html_content += """
        <h2>Summary</h2>
        <table class="summary-table">
"""
            for key, value in self.summary.items():
                if key != 'Generated':
                    html_content += f"            <tr><td><strong>{key}</strong></td><td>{self._format_value(value)}</td></tr>\n"
            html_content += "        </table>\n"

        # Data table
        if self.dataframe is not None and not self.dataframe.empty:
            html_content += """
        <h2>Details</h2>
        <table>
            <thead>
                <tr>
"""
            for col in self.dataframe.columns:
                html_content += f"                    <th>{col}</th>\n"
            html_content += """                </tr>
            </thead>
            <tbody>
"""
            for _, row in self.dataframe.iterrows():
                # Determine row class based on status
                status = str(row.get('Status', '')).lower()
                row_class = ''
                if status == 'present':
                    row_class = 'status-present'
                elif status == 'missing':
                    row_class = 'status-missing'
                elif status == 'extra':
                    row_class = 'status-extra'

                html_content += f'                <tr class="{row_class}">\n'
                for col in self.dataframe.columns:
                    value = self._format_value(row[col])
                    # Add class for match columns
                    cell_class = ''
                    if 'Match' in col and value == 'Yes':
                        cell_class = 'match-yes'
                    elif 'Match' in col and value == 'No':
                        cell_class = 'match-no'
                    elif 'Match' in col and value == 'Partial':
                        cell_class = 'match-partial'
                    html_content += f'                    <td class="{cell_class}">{value}</td>\n'
                html_content += '                </tr>\n'

            html_content += """            </tbody>
        </table>
"""

        html_content += """    </div>
</body>
</html>
"""
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)

        return filepath

    def save_pdf(self, filepath: Path) -> Path:
        """Export to PDF file (requires weasyprint or falls back to HTML)."""
        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)

        try:
            from weasyprint import HTML
            # Generate HTML first, then convert to PDF
            html_path = filepath.with_suffix('.html')
            self.save_html(html_path)
            HTML(filename=str(html_path)).write_pdf(str(filepath))
            html_path.unlink()  # Remove temp HTML
            return filepath
        except ImportError:
            # Fallback: save as HTML with .pdf.html extension
            print("Warning: weasyprint not installed. Install with: pip install weasyprint")
            print("Saving as HTML instead...")
            return self.save_html(filepath.with_suffix('.html'))

    def save(self, filepath: Path, format: str = 'md') -> Path:
        """
        Save report in specified format.

        Args:
            filepath: Output file path (extension will be adjusted)
            format: Output format - 'md', 'csv', 'xlsx', 'html', 'pdf'

        Returns:
            Path to saved file
        """
        filepath = Path(filepath)

        format_handlers = {
            'md': (self.save_markdown, '.md'),
            'markdown': (self.save_markdown, '.md'),
            'csv': (self.save_csv, '.csv'),
            'xlsx': (self.save_excel, '.xlsx'),
            'excel': (self.save_excel, '.xlsx'),
            'html': (self.save_html, '.html'),
            'pdf': (self.save_pdf, '.pdf'),
        }

        handler, extension = format_handlers.get(format.lower(), (self.save_markdown, '.md'))
        output_path = filepath.with_suffix(extension)

        return handler(output_path)

    def save_all(self, base_path: Path) -> Dict[str, Path]:
        """
        Save report in all available formats.

        Args:
            base_path: Base file path (without extension)

        Returns:
            Dict mapping format to saved file path
        """
        base_path = Path(base_path)
        results = {}

        for fmt in ['md', 'csv', 'xlsx', 'html']:
            try:
                results[fmt] = self.save(base_path, fmt)
            except Exception as e:
                print(f"Warning: Failed to save {fmt}: {e}")

        return results


def create_pod_check_report(
    summary: Dict[str, Any],
    dataframe: pd.DataFrame,
    title: str = "POD Check Report"
) -> MarkdownReport:
    """
    Create a POD check report.

    Args:
        summary: Summary statistics dict
        dataframe: Report data
        title: Report title

    Returns:
        MarkdownReport instance
    """
    report = MarkdownReport(title)
    report.set_summary(summary)
    report.set_data(dataframe)
    return report
